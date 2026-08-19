from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from uuid import UUID
from fastapi import HTTPException

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.core.base.services import Service
from app.utils.settings import settings

INTERNAL_KEY = settings.INTERNAL_KEY
BASE_URL = settings.INTERNAL_BASE_URL


def _to_number(value):
    """Coerce a numeric-looking value to a real number so Excel treats it as one.

    The internal API returns money amounts as strings (stuckValue is
    "11000705000"), which would otherwise land in the sheet as text and refuse
    to sum or format. Anything genuinely non-numeric is passed through as-is.
    """
    if value is None:
        return 0
    if isinstance(value, (int, float, Decimal)):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return value

class InternalAPIService(Service):
    
    @staticmethod
    def _request(method: str, endpoint: str, **kwargs):
        url = f"{BASE_URL}{endpoint}"
        headers = kwargs.pop("headers", {})
        headers.setdefault("x-internal-key", INTERNAL_KEY)
        
        # Clean up None values from params
        if "params" in kwargs and kwargs["params"]:
            kwargs["params"] = {k: v for k, v in kwargs["params"].items() if v is not None}
            
        response = requests.request(method, url, headers=headers, **kwargs)
        response.raise_for_status()
        return response.json()
    
    @staticmethod
    def search_users(query: str):
        return InternalAPIService._request("GET", "/internal/users/search", params={"q": query})
        
    @staticmethod
    def get_all_users(page: int = 1, limit: int = 10, **kwargs):
        params = {"page": page, "limit": limit, **kwargs}
        return InternalAPIService._request("GET", "/internal/users/all", params=params)

    @staticmethod
    def export_customers(**filters) -> bytes:
        """Build an .xlsx of every customer matching `filters` (same query params
        accepted by get_all_users), paging through the internal API to collect them all."""
        page = 1
        items = []
        while True:
            response = InternalAPIService.get_all_users(page=page, limit=500, **filters)
            data = response.get("data", {})
            items.extend(data.get("items", []))

            meta = data.get("meta", {})
            if page >= meta.get("totalPages", page):
                break
            page += 1

        wb = Workbook()
        ws = wb.active or wb.create_sheet(title="Customers")
        ws.title = "Customers"

        ws.append(
            ["Full Name", "Email", "Phone Number", "KYC Tier", "Status", "Date Joined"]
        )

        for item in items:
            full_name = " ".join(
                part for part in [item.get("firstName"), item.get("lastName")] if part
            ) or "-"

            created_at = item.get("createdAt")
            date_joined = None
            if created_at:
                date_joined = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                if date_joined.tzinfo is not None:
                    date_joined = date_joined.astimezone(timezone.utc).replace(tzinfo=None)

            ws.append(
                [
                    full_name,
                    item.get("email") or "-",
                    item.get("phoneNumber") or "-",
                    f"Tier {item.get('tier', 0)}",
                    "Active" if item.get("isActive") else "Inactive",
                    date_joined,
                ]
            )

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream.getvalue()
    
    @staticmethod
    def get_user_by_id(id: UUID):
        try:
            return InternalAPIService._request("GET", f"/internal/users/{id}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @staticmethod
    def log_manual_transaction(payload) -> dict:
        """Proxies the log manual transaction request to the core backend."""
        return InternalAPIService._request("POST", "/internal/wallet/transactions/manual", json=payload)

    @staticmethod
    def get_all_transactions(page: int = 1, limit: int = 10, **kwargs):
        params = {"page": page, "limit": limit, **kwargs}
        return InternalAPIService._request("GET", "/internal/wallet/transactions/all", params=params)

    @staticmethod
    def export_transactions(**filters) -> bytes:
        """Build an .xlsx of every transaction matching `filters` (same query params
        accepted by get_all_transactions), paging through the internal API to collect them all."""
        page = 1
        items = []
        while True:
            response = InternalAPIService.get_all_transactions(page=page, limit=500, **filters)
            data = response.get("data", {})
            items.extend(data.get("items", []))

            meta = data.get("meta", {})
            if page >= meta.get("totalPages", page):
                break
            page += 1

        wb = Workbook()
        ws = wb.active or wb.create_sheet(title="Transactions")
        ws.title = "Transactions"

        ws.append(
            ["Txn ID", "Date", "Customer", "Tier Level", "Details", "Amount", "Status"]
        )

        for item in items:
            customer = item.get("customer") or {}
            customer_name = " ".join(
                part for part in [customer.get("firstName"), customer.get("lastName")] if part
            ) or "-"

            kyc_tier = customer.get("kycTier")
            tier_level = f"Tier {kyc_tier}" if kyc_tier is not None else "-"

            details = item.get("displayType") or item.get("type") or "-"
            if item.get("note"):
                details = f"{details}\n{item['note']}"

            amount = item.get("amount")
            currency = item.get("currency")
            amount_display = "-"
            if amount is not None:
                formatted_amount = f"{float(amount):.2f}".rstrip("0").rstrip(".")
                amount_display = f"{currency} {formatted_amount}" if currency else formatted_amount

            created_at = item.get("createdAt")
            date = None
            if created_at:
                date = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                if date.tzinfo is not None:
                    date = date.astimezone(timezone.utc).replace(tzinfo=None)

            ws.append(
                [
                    item.get("reference") or "-",
                    date,
                    customer_name,
                    tier_level,
                    details,
                    amount_display,
                    item.get("status") or "-",
                ]
            )

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream.getvalue()

    @staticmethod
    def get_transaction_by_id(id: UUID):
        try:
            return InternalAPIService._request("GET", f"/internal/wallet/transactions/{id}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @staticmethod
    def get_user_transactions(user_id: UUID, page: int = 1, limit: int = 10, **kwargs):
        try:
            params = {"page": page, "limit": limit, **kwargs}
            return InternalAPIService._request("GET", f"/internal/users/{user_id}/transactions", params=params)
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
        
    @staticmethod
    def get_user_assets(user_id: UUID):
        try:
            return InternalAPIService._request("GET", f"/internal/users/{user_id}/assets")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
        
    @staticmethod
    def get_user_verification_details(user_id: UUID):
        try:
            return InternalAPIService._request("GET", f"/internal/users/{user_id}/verification")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
        

    @staticmethod
    def toggle_user_status(user_id: UUID):
        try:
            return InternalAPIService._request("PUT", f"/internal/users/{user_id}/status")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))


    @staticmethod
    def get_transaction_metrics():
        try:
            return InternalAPIService._request("GET", "/internal/metrics/transactions")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @staticmethod
    def get_customer_metrics():
        try:
            return InternalAPIService._request("GET", "/internal/metrics/customers")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @staticmethod
    def export_dashboard_metrics() -> bytes:
        """Build a single .xlsx report of the ERP dashboard metrics.

        One section per dashboard card, stacked down the sheet: a title row, a
        row of metric names, then a row of values. Metrics are a point-in-time
        summary rather than a row-per-entity list, so the report closes with a
        generation timestamp — without it a downloaded snapshot can't be placed
        after the fact.
        """
        transactions = InternalAPIService.get_transaction_metrics().get("data") or {}
        customers = InternalAPIService.get_customer_metrics().get("data") or {}

        sections = [
            (
                "Transactions",
                [
                    ("Total", transactions.get("total", 0)),
                    ("Success", transactions.get("success", 0)),
                    ("Pending", transactions.get("pending", 0)),
                    ("Failed", transactions.get("failed", 0)),
                    ("Stuck in Pending", transactions.get("stuckInPending", 0)),
                ],
            ),
            (
                "Customers",
                [
                    ("Total", customers.get("total", 0)),
                    ("New Today", customers.get("newToday", 0)),
                    ("KYC Pending", customers.get("kycPending", 0)),
                    ("Flagged", customers.get("flagged", 0)),
                ],
            ),
        ]

        wb = Workbook()
        ws = wb.active or wb.create_sheet()
        ws.title = "Dashboard Metrics"

        bold = Font(bold=True)

        for title, metrics in sections:
            ws.append([title])
            ws.cell(row=ws.max_row, column=1).font = bold

            ws.append([label for label, _ in metrics])
            for column in range(1, len(metrics) + 1):
                ws.cell(row=ws.max_row, column=column).font = bold

            ws.append([value for _, value in metrics])
            ws.append([])

        ws.append(["Generated At (UTC)", datetime.now(timezone.utc).replace(tzinfo=None)])
        ws.cell(row=ws.max_row, column=1).font = bold

        # widest label across every section, so nothing is truncated
        widest = max(len(label) for _, metrics in sections for label, _ in metrics)
        for column in range(1, max(len(m) for _, m in sections) + 1):
            ws.column_dimensions[get_column_letter(column)].width = widest + 4

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream.getvalue()

    @staticmethod
    def get_user_kyc_details(user_id: UUID):
        try:
            return InternalAPIService._request("GET", f"/internal/users/{user_id}/kyc")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))