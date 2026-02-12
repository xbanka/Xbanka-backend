import logging
import secrets
from app.core.base.services import Service
from app.core.enums import PayoutStatusEnum
from app.core.hash import Hasher
from app.models.affiliate import Affiliate
from app.models.affiliate_commissions import AffiliateCommission
from app.models.affiliate_tiers import AffiliateTier
from app.models.affiliate_visit import AffiliateVisit
from app.models.payouts import Payout
from app.models.transactions import Transaction
from app.models.customer import Customer
from app.schemas.affiliate import UpdateBankDetailsRequest
from app.schemas.user import RegisterBase
from app.utils.validators import is_valid_account_number, is_valid_email, is_valid_password, is_valid_phone
from app.utils.visitors import generate_visitor_id, already_counted_today, generate_fingerprint

from datetime import datetime, timezone
from fastapi import HTTPException, Request, Response, status
from io import BytesIO
from sqlalchemy import select, func, or_
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy.orm import Session, selectinload
from openpyxl import Workbook
from typing import Optional, cast


logger = logging.getLogger(__name__)


class AffiliateService(Service):
    @staticmethod
    def create(db: Session, obj_in: RegisterBase) -> Affiliate:

        if not is_valid_email(obj_in.email):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='Invalid email address.'
            )

        stmt = select(Affiliate).where(
            or_(
                Affiliate.email == obj_in.email,
                Affiliate.username == obj_in.username
            )
        )
        affiliate = db.execute(stmt).scalars().first()

        if affiliate and not affiliate.verified:
            return affiliate

        if affiliate:
            raise HTTPException(
                status_code=400,
                detail="Affiliate with email/username already exists"
            )
    
        if not is_valid_password(obj_in.password):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='Password must be at least 8 characters long and contain at least one uppercase letter, one lowercase letter, one digit, and one special character.'
            )
        
        if not is_valid_phone(obj_in.phone_no):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='Phone number must be a valid Nigerian (+234) or international format'
            )

        while True:
            generated_code = secrets.token_urlsafe(16)
            if not db.query(Affiliate).filter_by(ref_code=generated_code).first():
                break

        bronze_tier = db.execute(
            select(AffiliateTier).where(AffiliateTier.name == "Bronze")
        ).scalars().first()

        try:
            affiliate = Affiliate(
                first_name=obj_in.first_name,
                last_name=obj_in.last_name,
                email=obj_in.email,
                username=obj_in.username,
                phone_no=obj_in.phone_no,
                bank=obj_in.bank,
                account_no=obj_in.account_no,
                hashed_password=Hasher.get_password_hash(obj_in.password),
                verified=False,
                ref_code=generated_code,
                current_tier_id=bronze_tier.id if bronze_tier else None
            )

            db.add(affiliate)
            db.commit()
            db.refresh(affiliate)
        except IntegrityError:
            logger.exception("Integrity error while creating affiliate")
            db.rollback()
            raise HTTPException(status_code=400, detail="Database integrity error")
        except SQLAlchemyError:
            logger.exception("SQLAlchemy error while creating affiliate")
            db.rollback()
            raise HTTPException(status_code=500, detail="An error occurred saving entity")
        except Exception as e:
            logger.exception("Unexpected error while creating affiliate")
            db.rollback()
            raise HTTPException(status_code=500, detail="An unknown error occurred")

        return affiliate

    @staticmethod
    def get_user_by_id(db: Session, id: str) -> Affiliate:
        affiliate = db.query(Affiliate).get(id)
        if not affiliate:
            raise HTTPException(
                status_code=404,
                detail="Affiliate not found"
            )
        
        return affiliate

    @staticmethod
    def get_by_username(db: Session, username: str) -> Affiliate:
        stmt = select(Affiliate).where(Affiliate.username == username)
        affiliate = db.execute(stmt).scalars().one_or_none()
        if not affiliate:
            raise HTTPException(
                status_code=404,
                detail="Affiliate with this username not found"
            )
        return affiliate
    
    @staticmethod
    def get_by_refcode(db: Session, refcode: str) -> Affiliate:
        stmt = select(Affiliate).where(Affiliate.ref_code == refcode)
        affiliate = db.execute(stmt).scalars().one_or_none()
        if not affiliate:
            raise HTTPException(
                status_code=404,
                detail="Affiliate with this referral code not found"
            )
        return affiliate
    
    @staticmethod
    def get_user_by_mail(db: Session, email: str) -> Affiliate:
        affiliate = db.query(Affiliate).filter_by(email=email).first()
        if not affiliate:
            raise HTTPException(
                status_code=404,
                detail="affiliate not found"
            )
        
        return affiliate
    
    @staticmethod
    def set_codename(db: Session, affiliate: Affiliate, codename: str) -> Affiliate:
        # check if codename already exists
        existing = db.query(Affiliate).filter_by(custom_refcode=codename).first()
        if existing and existing.id == affiliate.id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Codename already in use"
            )
        
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Codename already taken"
            )

        # set codename
        affiliate.custom_refcode = codename
        db.commit()
        db.refresh(affiliate)

        return affiliate
    
    @staticmethod
    def get_referral_code(db: Session, current_user: Affiliate):
        affiliate: Affiliate | None = db.query(Affiliate).get(current_user.id)
        if not affiliate:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Affiliate not found"
            )
        return affiliate.ref_code
    
    @staticmethod
    def record_unique_visit(db: Session, affiliate: Affiliate, request: Request, response: Response) -> None:

        try:
            # 1. Visitor ID from cookie
            visitor_id = request.cookies.get("affiliate_visitor_id")

            if not visitor_id:
                visitor_id = generate_visitor_id()
                response.set_cookie(
                    key="affiliate_visitor_id",
                    value=visitor_id,
                    httponly=True,
                    max_age=60 * 60 * 24 * 365,  # 1 year
                    samesite="lax",
                )
            
            # 2. Fingerprint fallback
            if request.client is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Unable to determine client IP address"
                )
            
            ip = request.client.host
            user_agent = request.headers.get("user-agent", "")
            fingerprint = generate_fingerprint(ip, user_agent)

            # 3. Deduplicate
            if not already_counted_today(
                db=db,
                affiliate_id=affiliate.id,
                visitor_id=visitor_id,
                fingerprint=fingerprint,
            ):
                visit = AffiliateVisit(
                    affiliate_id=affiliate.id,
                    visitor_id=visitor_id,
                    fingerprint=fingerprint,
                )
                db.add(visit)
                db.commit()

            # redirect to frontend?

        except Exception as e:
            logger.exception("Failed to record affiliate visit")
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to record affiliate visit"
            ) from e

    @staticmethod
    def get_all(db: Session):
        all_affiliates = db.query(Affiliate).all()
        return all_affiliates
    

    @staticmethod
    def get_commissions(db: Session, affiliate_id, page: int, limit: int):
        stmt = select(AffiliateCommission).join(Transaction)

        stmt = (
            stmt.options(selectinload(AffiliateCommission.transaction).selectinload(Transaction.customer))
            .where(AffiliateCommission.affiliate_id == affiliate_id)
            .order_by(AffiliateCommission.created_at.desc())
            .offset((page - 1) * limit)
            .limit(limit)
        )

        result = db.execute(stmt)
        commissions = result.scalars().all()

        total = db.scalar(
            select(func.count()).select_from(
                select(AffiliateCommission)
                .where(AffiliateCommission.affiliate_id == affiliate_id)
                .subquery()
            )
        ) or 0

        return {
            "page": page,
            "limit": limit,
            "total": total,
            "pages": (total + limit - 1) // limit,
            "data": commissions
        }
    
    @staticmethod
    def export_commissions(db: Session, affiliate_id):
        stmt = select(AffiliateCommission).join(Transaction)

        stmt = (
            stmt.options(selectinload(AffiliateCommission.transaction).selectinload(Transaction.customer))
            .where(AffiliateCommission.affiliate_id == affiliate_id)
            .order_by(AffiliateCommission.created_at.desc())
        )

        result = db.execute(stmt)
        commissions = result.scalars().all()

        wb = Workbook()
        ws = wb.active or wb.create_sheet(title="Commissions")
        ws.title = "Commissions"

        ws.append(
            ["Date", "Type", "Rate", "Amount", "Status", "Customer Name", "Email", "Phone No"]
        )  # header
        
        for tx in commissions:
            # Excel does not support timezone-aware datetime objects
            created_at = cast(datetime, tx.created_at)
            if created_at.tzinfo is not None:
                created_at = created_at.astimezone(timezone.utc).replace(tzinfo=None)

            ws.append([
                created_at,
                tx.transaction_type,
                tx.commission_rate,
                tx.commission_amount,
                tx.status.value,
                tx.customer.first_name + " " + tx.customer.last_name,
                tx.customer.email,
                tx.customer.phone_no,
            ])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream.getvalue()

    @staticmethod
    def get_payouts(db: Session, affiliate_id, page: int, limit: int, status: Optional[PayoutStatusEnum] = None):
        stmt = select(Payout)

        if status:
            stmt = stmt.where(Payout.status == status)

        stmt = (
            stmt.where(Payout.affiliate_id == affiliate_id)
            .order_by(Payout.created_at.desc())
            .offset((page - 1) * limit)
            .limit(limit)
        )

        result = db.execute(stmt)
        payouts = result.scalars().all()

        total = db.scalar(
            select(func.count()).select_from(
                select(Payout)
                .where(Payout.affiliate_id == affiliate_id)
                .subquery()
            )
        ) or 0

        return {
            "page": page,
            "limit": limit,
            "total": total,
            "pages": (total + limit - 1) // limit,
            "data": payouts
        }
    
    @staticmethod
    def export_payouts(db: Session, affiliate_id, status: Optional[PayoutStatusEnum] = None):
        stmt = select(Payout)

        if status:
            stmt = stmt.where(Payout.status == status)

        stmt = (
            stmt.where(Payout.affiliate_id == affiliate_id)
            .order_by(Payout.created_at.desc())
        )

        payouts = db.execute(stmt).scalars().all()

        wb = Workbook()
        ws = wb.active or wb.create_sheet(title="Payouts")
        ws.title = "Payouts"

        ws.append(
            ["Amount", "Status", "Payment Reference", "Paid At"]
        )  # header
        
        for payout in payouts:
            # Excel does not support timezone-aware datetime objects
            created_at = cast(datetime, payout.created_at)
            if created_at.tzinfo is not None:
                created_at = created_at.astimezone(timezone.utc).replace(tzinfo=None)

            paid_at = cast(datetime, payout.paid_at)
            if paid_at.tzinfo is not None:
                paid_at = paid_at.astimezone(timezone.utc).replace(tzinfo=None)

            ws.append([
                payout.amount,
                payout.status.value,
                payout.payment_ref,
                paid_at
            ])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream.getvalue()
    

    @staticmethod
    def get_payouts_summary(db: Session, affiliate_id):
        total_earnings = db.scalar(
            select(func.coalesce(func.sum(Transaction.commission_amount), 0))
            .join(Customer, Customer.id == Transaction.customer_id)
            .where(Customer.affiliate_id == affiliate_id)
        ) or 0

        total_payouts = db.scalar(
            select(func.coalesce(func.sum(Payout.amount), 0))
            .where(Payout.affiliate_id == affiliate_id)
            .where(Payout.status == PayoutStatusEnum.paid)
        ) or 0

        amount_withdrawn = total_payouts
        available_payouts = total_earnings - amount_withdrawn

        pending_payouts = db.scalar(
            select(func.coalesce(func.sum(Payout.amount), 0))
            .where(Payout.affiliate_id == affiliate_id)
            .where(Payout.status == PayoutStatusEnum.pending)
        ) or 0

        failed_payouts = db.scalar(
            select(func.coalesce(func.sum(Payout.amount), 0))
            .where(Payout.affiliate_id == affiliate_id)
            .where(Payout.status == PayoutStatusEnum.failed)
        ) or 0

        pending_earnings = db.scalar(
            select(func.coalesce(func.sum(Transaction.commission_amount), 0))
            .join(Customer, Customer.id == Transaction.customer_id)
            .where(Customer.affiliate_id == affiliate_id)
        ) or 0

        return {
            "total_earnings": total_earnings,
            "pending_payouts": pending_payouts,
            "amount_withdrawn": amount_withdrawn,
            "available_balance": available_payouts,
        }
    

    @staticmethod
    def update_bank_details(
        db: Session, 
        affiliate: Affiliate, 
        bank_details: UpdateBankDetailsRequest
    ) -> None:
        
        # Validate Nigerian bank account number (10 digits)
        if not is_valid_account_number(bank_details.account_number):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid account number"
            )
        
        try:
            affiliate.bank = bank_details.bank_name
            affiliate.account_no = bank_details.account_number
            db.commit()
        except Exception as e:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to update bank details"
            ) from e
