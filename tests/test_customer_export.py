from io import BytesIO

from openpyxl import load_workbook


def _page(items, total_pages=1, current_page=1):
    return {
        "message": "Request successful",
        "data": {
            "items": items,
            "meta": {
                "totalItems": len(items),
                "itemCount": len(items),
                "itemsPerPage": 500,
                "totalPages": total_pages,
                "currentPage": current_page,
            },
        },
    }


CUSTOMER_NO_NAME = {
    "userId": "712a2800-f840-4848-bd6c-3ec1a510e2d8",
    "email": "garrygrado@paldonanamansilolo.online",
    "isActive": True,
    "firstName": None,
    "lastName": None,
    "phoneNumber": None,
    "tier": 0,
    "createdAt": "2026-08-05T16:26:09.259Z",
}

CUSTOMER_WITH_NAME = {
    "userId": "7b53280d-a7f9-4997-8e1a-37387ddd46da",
    "email": "eyebiokinjoseph1@gmail.com",
    "isActive": True,
    "firstName": "Joseph",
    "lastName": "Eyebiokin",
    "phoneNumber": "07089462051",
    "tier": 2,
    "createdAt": "2026-08-05T11:16:58.597Z",
}

CUSTOMER_INACTIVE = {
    "userId": "9c1a2800-f111-4848-bd6c-3ec1a510e2d9",
    "email": "inactive@example.com",
    "isActive": False,
    "firstName": "Sam",
    "lastName": None,
    "phoneNumber": None,
    "tier": 1,
    "createdAt": "2026-07-01T09:00:00.000Z",
}


def _load_rows(response_content):
    wb = load_workbook(BytesIO(response_content))
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def test_export_customers_success(mocker, super_client):
    mock_get_all_users = mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_all_users",
        return_value=_page([CUSTOMER_NO_NAME, CUSTOMER_WITH_NAME, CUSTOMER_INACTIVE]),
    )

    response = super_client.get("/api/customers/export")

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "customers.xlsx" in response.headers["content-disposition"]

    rows = _load_rows(response.content)
    assert rows[0] == ("Full Name", "Email", "Phone Number", "KYC Tier", "Status", "Date Joined")
    assert rows[1][0] == "-"  # no first/last name
    assert rows[1][2] == "-"  # no phone
    assert rows[1][3] == "Tier 0"
    assert rows[1][4] == "Active"

    assert rows[2][0] == "Joseph Eyebiokin"
    assert rows[2][2] == "07089462051"
    assert rows[2][3] == "Tier 2"

    assert rows[3][4] == "Inactive"

    mock_get_all_users.assert_called_once_with(
        page=1, limit=500, search=None, status=None, startDate=None, endDate=None
    )


def test_export_customers_passes_filters_through(mocker, super_client):
    mock_get_all_users = mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_all_users",
        return_value=_page([CUSTOMER_WITH_NAME]),
    )

    response = super_client.get(
        "/api/customers/export",
        params={
            "search": "joseph",
            "status": "ACTIVE",
            "startDate": "2026-01-01",
            "endDate": "2026-08-01",
        },
    )

    assert response.status_code == 200
    mock_get_all_users.assert_called_once_with(
        page=1,
        limit=500,
        search="joseph",
        status="ACTIVE",
        startDate="2026-01-01",
        endDate="2026-08-01",
    )


def test_export_customers_paginates_until_exhausted(mocker, super_client):
    mock_get_all_users = mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_all_users",
        side_effect=[
            _page([CUSTOMER_NO_NAME], total_pages=2, current_page=1),
            _page([CUSTOMER_WITH_NAME], total_pages=2, current_page=2),
        ],
    )

    response = super_client.get("/api/customers/export")

    assert response.status_code == 200
    rows = _load_rows(response.content)
    assert len(rows) == 3  # header + one row per page

    assert mock_get_all_users.call_count == 2
    first_call, second_call = mock_get_all_users.call_args_list
    assert first_call.kwargs["page"] == 1
    assert second_call.kwargs["page"] == 2
    assert first_call.kwargs["limit"] == 500


def test_export_customers_no_results_returns_header_only(mocker, super_client):
    mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_all_users",
        return_value=_page([]),
    )

    response = super_client.get("/api/customers/export")

    assert response.status_code == 200
    rows = _load_rows(response.content)
    assert len(rows) == 1
    assert rows[0] == ("Full Name", "Email", "Phone Number", "KYC Tier", "Status", "Date Joined")


def test_export_customers_upstream_failure_returns_500(mocker, super_client):
    mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_all_users",
        side_effect=RuntimeError("internal backend unreachable"),
    )

    response = super_client.get("/api/customers/export")

    assert response.status_code == 500
