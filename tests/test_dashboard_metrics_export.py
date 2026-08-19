from io import BytesIO

from openpyxl import load_workbook


TRANSACTION_METRICS = {
    "message": "Request successful",
    "details": "",
    "errorGroup": "",
    "data": {
        "total": 412,
        "success": 395,
        "pending": 7,
        "failed": 10,
        "stuckInPending": 7,
        "stuckValue": "11000705000",
    },
}

CUSTOMER_METRICS = {
    "message": "Request successful",
    "details": "",
    "errorGroup": "",
    "data": {
        "total": 130,
        "newToday": 0,
        "kycPending": 21,
        "flagged": 0,
    },
}

URL = "/api/dashboard/erp/metrics/export"


def _patch_metrics(mocker, transactions=None, customers=None):
    mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_transaction_metrics",
        return_value=transactions if transactions is not None else TRANSACTION_METRICS,
    )
    mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_customer_metrics",
        return_value=customers if customers is not None else CUSTOMER_METRICS,
    )


def _load_rows(response_content):
    wb = load_workbook(BytesIO(response_content))
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def test_export_metrics_success(mocker, super_client):
    _patch_metrics(mocker)

    response = super_client.get(URL)

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "dashboard-metrics.xlsx" in response.headers["content-disposition"]


def test_export_metrics_lays_each_card_out_as_its_own_section(mocker, super_client):
    _patch_metrics(mocker)

    rows = _load_rows(super_client.get(URL).content)

    assert rows[0][0] == "Transactions"
    assert rows[1][:5] == (
        "Total",
        "Success",
        "Pending",
        "Failed",
        "Stuck in Pending"
    )
    assert rows[2][:5] == (412, 395, 7, 10, 7)

    assert rows[4][0] == "Customers"
    assert rows[5][:4] == ("Total", "New Today", "KYC Pending", "Flagged")
    assert rows[6][:4] == (130, 0, 21, 0)


def test_export_metrics_writes_stuck_value_as_a_number(mocker, super_client):
    """stuckValue arrives as a string; as text Excel won't sum or format it."""
    _patch_metrics(mocker)

    rows = _load_rows(super_client.get(URL).content)

    assert isinstance(rows[2][4], (int, float))


def test_export_metrics_includes_generation_timestamp(mocker, super_client):
    from datetime import datetime

    _patch_metrics(mocker)

    rows = _load_rows(super_client.get(URL).content)
    label, value = rows[-1][0], rows[-1][1]

    assert label == "Generated At (UTC)"
    assert isinstance(value, datetime)


def test_export_metrics_tolerates_missing_keys(mocker, super_client):
    """A partial upstream payload should still produce a report, not a 500."""
    _patch_metrics(
        mocker,
        transactions={"data": {"total": 5}},
        customers={"data": {}},
    )

    response = super_client.get(URL)

    assert response.status_code == 200
    rows = _load_rows(response.content)
    assert rows[2][:5] == (5, 0, 0, 0, 0)
    assert rows[6][:4] == (0, 0, 0, 0)


def test_export_metrics_upstream_failure_returns_500(mocker, super_client):
    mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_transaction_metrics",
        side_effect=RuntimeError("internal backend unreachable"),
    )

    response = super_client.get(URL)

    assert response.status_code == 500


def test_export_metrics_requires_auth(test_client):
    response = test_client.get(URL)
    assert response.status_code == 401
