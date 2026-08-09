from io import BytesIO

from openpyxl import load_workbook


def _page(items, total_pages=1, current_page=1):
    return {
        "message": "Request successful",
        "data": {
            "items": items,
            "meta": {
                "totalItems": len(items),
                "itemCount": len(items),
                "itemsPerPage": 500,
                "totalPages": total_pages,
                "currentPage": current_page,
            },
        },
    }


TXN_WITH_CUSTOMER = {
    "id": "c4ec8850-f93e-40e3-8bb8-8bbdf0a2452b",
    "reference": "CONVC-1786302942185",
    "createdAt": "2026-08-09T19:15:42.186Z",
    "status": "COMPLETED",
    "type": "DEPOSIT",
    "displayType": "Crypto Deposit",
    "currency": "USDT",
    "amount": 0.3597,
    "note": "Conversion Credit: Received 0.3597 USDT from NGN",
    "category": "CRYPTO",
    "customer": {
        "id": "991fbf45-8c5a-4840-9f4f-8af81c586922",
        "firstName": "Ayodeji",
        "lastName": "Adebayo",
        "email": "dejibayo0630@gmail.com",
        "phoneNumber": "8064180322",
        "kycTier": 1,
    },
}

TXN_NO_CUSTOMER = {
    "id": "500e1ca2-3e85-4244-8a53-0b7330c0a41d",
    "reference": "CONVD-1786302920189",
    "createdAt": "2026-07-01T09:00:00.000Z",
    "status": "PENDING",
    "type": "WITHDRAWAL",
    "displayType": None,
    "currency": "NGN",
    "amount": 500,
    "note": None,
    "category": "FIAT",
    "customer": None,
}


def _load_rows(response_content):
    wb = load_workbook(BytesIO(response_content))
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def test_export_transactions_success(mocker, super_client):
    mock_get_all_transactions = mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_all_transactions",
        return_value=_page([TXN_WITH_CUSTOMER, TXN_NO_CUSTOMER]),
    )

    response = super_client.get("/api/transactions/export")

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "transactions.xlsx" in response.headers["content-disposition"]

    rows = _load_rows(response.content)
    assert rows[0] == ("Txn ID", "Date", "Customer", "Tier Level", "Details", "Amount", "Status")

    assert rows[1][0] == "CONVC-1786302942185"
    assert rows[1][2] == "Ayodeji Adebayo"
    assert rows[1][3] == "Tier 1"
    assert rows[1][4] == "Crypto Deposit\nConversion Credit: Received 0.3597 USDT from NGN"
    assert rows[1][5] == "USDT 0.36"
    assert rows[1][6] == "COMPLETED"

    # no customer on the record -> falls back to "-", displayType falls back to
    # the raw type, whole-number amount doesn't carry trailing decimal zeros
    assert rows[2][2] == "-"
    assert rows[2][3] == "-"
    assert rows[2][4] == "WITHDRAWAL"
    assert rows[2][5] == "NGN 500"

    mock_get_all_transactions.assert_called_once_with(
        page=1,
        limit=500,
        search=None,
        status=None,
        type=None,
        currency=None,
        category=None,
        startDate=None,
        endDate=None,
    )


def test_export_transactions_passes_filters_through(mocker, super_client):
    mock_get_all_transactions = mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_all_transactions",
        return_value=_page([TXN_WITH_CUSTOMER]),
    )

    response = super_client.get(
        "/api/transactions/export",
        params={
            "search": "ayodeji",
            "status": "COMPLETED",
            "type": "DEPOSIT",
            "currency": "USDT",
            "category": "CRYPTO",
            "startDate": "2026-01-01",
            "endDate": "2026-08-01",
        },
    )

    assert response.status_code == 200
    mock_get_all_transactions.assert_called_once_with(
        page=1,
        limit=500,
        search="ayodeji",
        status="COMPLETED",
        type="DEPOSIT",
        currency="USDT",
        category="CRYPTO",
        startDate="2026-01-01",
        endDate="2026-08-01",
    )


def test_export_transactions_paginates_until_exhausted(mocker, super_client):
    mock_get_all_transactions = mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_all_transactions",
        side_effect=[
            _page([TXN_WITH_CUSTOMER], total_pages=2, current_page=1),
            _page([TXN_NO_CUSTOMER], total_pages=2, current_page=2),
        ],
    )

    response = super_client.get("/api/transactions/export")

    assert response.status_code == 200
    rows = _load_rows(response.content)
    assert len(rows) == 3  # header + one row per page

    assert mock_get_all_transactions.call_count == 2
    first_call, second_call = mock_get_all_transactions.call_args_list
    assert first_call.kwargs["page"] == 1
    assert second_call.kwargs["page"] == 2
    assert first_call.kwargs["limit"] == 500


def test_export_transactions_no_results_returns_header_only(mocker, super_client):
    mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_all_transactions",
        return_value=_page([]),
    )

    response = super_client.get("/api/transactions/export")

    assert response.status_code == 200
    rows = _load_rows(response.content)
    assert len(rows) == 1


def test_export_transactions_upstream_failure_returns_500(mocker, super_client):
    mocker.patch(
        "app.services.internal_backend.InternalAPIService.get_all_transactions",
        side_effect=RuntimeError("internal backend unreachable"),
    )

    response = super_client.get("/api/transactions/export")

    assert response.status_code == 500


def test_export_transactions_requires_auth(test_client):
    response = test_client.get("/api/transactions/export")
    assert response.status_code == 401
